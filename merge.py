import pandas as pd
from openpyxl import load_workbook

# Function to handle merged cells and take only the first word in the second column
def unmerge_and_fill_first_word(excel_file, sheet_name):
    # Load the workbook and the specific sheet
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Copy the merged cells ranges to a list (to avoid set size change issue)
    merged_cells_ranges = list(ws.merged_cells.ranges)
    
    # Iterate through the copied list of merged cell ranges
    for merged_cell in merged_cells_ranges:
        # Unmerge each cell
        ws.unmerge_cells(str(merged_cell))
        
        # Get the top-left cell value of the merged range
        top_left_cell_value = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col).value
        
        # If the merged range includes the second column (e.g., column index = 2)
        if merged_cell.min_col == 2:  # Assuming the second column is at index 2 (column B)
            # Extract the first word of the top-left cell value
            first_word = str(top_left_cell_value).split()[0] if pd.notna(top_left_cell_value) else None
            
            # Fill the unmerged cells with the first word in the second column
            for row in ws.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                                    min_col=merged_cell.min_col, max_col=merged_cell.max_col):
                for cell in row:
                    cell.value = first_word
        else:
            # Fill other columns (if needed) with the top-left value or leave unchanged
            for row in ws.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                                    min_col=merged_cell.min_col, max_col=merged_cell.max_col):
                for cell in row:
                    if cell.row == merged_cell.min_row and cell.column == merged_cell.min_col:
                        cell.value = top_left_cell_value
                    else:
                        cell.value = None  # Leave others empty

    # Save the modified workbook after unmerging and filling cells
    modified_file = 'unmerged_filled_first_word.xlsx'
    wb.save(modified_file)
    
    # Load the modified Excel into a pandas DataFrame
    df = pd.read_excel(modified_file, sheet_name=sheet_name)
    
    # Save the modified DataFrame back to an Excel file if needed
    df.to_excel('unmerged_filled_first_word_processed.xlsx', index=False)
    print(f"Unmerged and filled cells saved to: 'unmerged_filled_first_word_processed.xlsx'")
    
    return df

# Example usage
excel_file = 'your_excel_file.xlsx'
sheet_name = 'Sheet1'

# Process the Excel sheet
df = unmerge_and_fill_first_word(excel_file, sheet_name)

# Display the modified DataFrame
print(df)
